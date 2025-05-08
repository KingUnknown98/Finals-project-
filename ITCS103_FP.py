
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from tkinter import PhotoImage
import os

def Sales():
    sales_w=tk.Toplevel(background="green4")
    sales_w.geometry("850x480")
    sales_w.resizable(height= False, width= False)
    sales_w.title("Sales")

    sbg=PhotoImage(file=r'D:\ITCS103_BOCA_1B\FinalsActivity\benta.png')
    tk.Label(sales_w,image=sbg).pack()

    tk.Button(sales_w, text = "Products Overview", width=15,bd=0,fg = "White",bg="blue4",font=("arial",13),command=Products).place(relx=.45,rely=.004)
    tk.Button(sales_w, text = "Sales Overview", width=15,bd=0,fg = "White",bg="medium blue",font=("arial",13)).place(relx=.62,rely=.004)
    tk.Button(sales_w, text = "Logout", width=15,bd=0,fg = "White",bg="#3533cd",font=("arial",13),command=sales_w.destroy).place(relx=.80,rely=.004)


    sb=Entry(sales_w, width=15,fg = "black",bg="White",font=("arial",15),bd=0)

    sb.place(width=200,height=30,rely=.10,relx=.70)


    tk.Button(sales_w, text = "Save",bd=0,fg = "black",bg="white",font=("arial",15,"bold")).place(relx=.15,rely=.90)
    tk.Button(sales_w, text = "Calculator",bd=0,fg = "black",bg="white",font=("arial",15,"bold")).place(relx=.42,rely=.89)
    tk.Button(sales_w, text = "Scratch",bd=0,fg = "black",bg="White",font=("arial",15,"bold")).place(relx=.73,rely=.90)

    sales_w.mainloop()



def Products():
    product_w=tk.Toplevel()
    product_w.geometry("850x480")
    product_w.resizable(width=False,height=False)
    product_w.title("Products")

    pbg=PhotoImage(file=r'D:\ITCS103_BOCA_1B\FinalsActivity\prod.png')
    tk.Label(product_w,image=pbg).pack()
    
    tk.Button(product_w, text = "Products Overview", width=15,bd=0,fg = "White",bg="blue4",font=("arial",13)).place(relx=.45,rely=.004)
    tk.Button(product_w, text = "Sales Overview", width=15,bd=0,fg = "White",bg="medium blue",font=("arial",13),command=Sales).place(relx=.62,rely=.004)
    tk.Button(product_w, text = "Logout", width=15,bd=0,fg = "White",bg="#3533cd",font=("arial",13),command=product_w.destroy).place(relx=.80,rely=.004)

    tk.Label(product_w,text="Products Overview",foreground="White",bg="black",font=("arial", 30)).place(rely=.03,relx=.03)
    tk.Button(product_w, text="Add",font=("arial", 16),foreground="white", bg="black",bd=0,width=9).place(rely=.79,relx=.61)
    tk.Button(product_w, text="Remove",font=("arial", 16),foreground="white", bg="black",bd=0,width=9).place(rely=.79,relx=.80)
    tk.Button(product_w, text="Update",font=("arial", 16),foreground="white", bg="black",bd=0,width=9).place(rely=.90,relx=.72)

    search=Entry(product_w,font=("arial",15),fg="black",bg="white",bd=0)
    search.place(width=130,rely=.10,relx=.80)
    
    p_name=Entry(product_w,font=("arial",15),fg="black",bg="white",bd=0)
    p_name.place(height=23,width=175,rely=.28,relx=.60)

    price=Entry(product_w,font=("arial",15),fg="black",bg="white",bd=0)
    price.place(height=23,width=175,rely=.41,relx=.60)

    expense=Entry(product_w,font=("arial",15),fg="black",bg="white",bd=0)
    expense.place(height=23,width=175,rely=.54,relx=.60)

    p_sold=Entry(product_w,font=("arial",15),fg="black",bg="white",bd=0)
    p_sold.place(height=23,width=175,rely=.68,relx=.60)

    product_w.mainloop()

def main_window():
    windows=tk.Tk()
    windows.geometry("850x480")
    windows.resizable(height=False,width=False)
    windows.title("Sales Information Software")

    home=PhotoImage(file=r'D:\ITCS103_BOCA_1B\FinalsActivity\homepage.png')
    tk.Label(windows,image=home).pack()

    l1=Label(windows,text="Dashboad",foreground="White",bg="black",font=("arial",30),bd=0)
    l1.place(relx=.02,rely=.02)

    search=Entry(windows,font=("arial",20),fg="black",bd=0)
    search.place(rely=.165,relx=.09,width=200)

    tk.Button(windows, text = "Products Overview", width=15,bd=0,fg = "White",bg="blue4",font=("arial",13),command=Products).place(relx=.45,rely=.004)
    tk.Button(windows, text = "Sales Overview", width=15,bd=0,fg = "White",bg="medium blue",font=("arial",13),command=Sales).place(relx=.62,rely=.004)
    tk.Button(windows, text = "Logout", width=15,bd=0,fg = "White",bg="#3533cd",font=("arial",13),command=windows.destroy).place(relx=.80,rely=.004)

    windows.mainloop()

file_path = "users.xlsx"
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Username", "Password"])
    wb.save(file_path)


wb = load_workbook(file_path)
ws = wb.active

def user_exists(username):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return True
    return False

def get_password(username):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return row[1]
    return None

def add_user(username, password):
    if not user_exists(username):
        ws.append([username, password])
        wb.save(file_path)
        return True
    return False

def login():
    username = username_entry.get()
    password = password_entry.get()
    if get_password(username) == password:
        messagebox.showinfo("Success", "Login Successful!")
        window.destroy()
        main_window()
    else:
        messagebox.showerror("Error", "Invalid username or password.")
    

def create_account():
    username = username_entry.get()
    password = password_entry.get()
    if username == "" or password == "":
        messagebox.showwarning("Warning", "Username and password required.")
    elif add_user(username, password):
        messagebox.showinfo("Success", "Account created successfully.")
    else:
        messagebox.showerror("Error", "User already exists.")
    
    username_entry.delete(0,tk.END)
    password_entry.delete(0,tk.END)

def forgot_password():
    username = username_entry.get()
    if not username:
        messagebox.showwarning("Warning", "Enter your username to retrieve password.")
    else:
        password = get_password(username)
        if password:
            messagebox.showinfo("Password Found", f"Password: {password}")
        else:
            messagebox.showerror("Error", "User not found.")


window= tk.Tk()
window.title("MIS")
window.config(bg="black")
window.geometry("850x480")

backgroundp=PhotoImage(file=r'D:\ITCS103_BOCA_1B\FinalsActivity\loginn.png')
tk.Label(window,image=backgroundp).pack()

tk.Label(window, text="Username",font=("Times",25),bg="Black",fg="White").place(rely=.27,relx=.58)
username_entry = tk.Entry(window,font=("Arial",20),bd=0)
username_entry.place(height=40,width=165,rely=.37,relx=.60)

tk.Label(window, text="Password",font=("times",25),bg="black",fg="White").place(rely=.48,relx=.58)
password_entry = tk.Entry(window, show="*",font=("Times",20),bd=0)
password_entry.place(width=165,height=40,rely=.58,relx=.60)

tk.Button(window, text="Login",font=("times",25), command=login,bg="Blue",fg="white",bd=0,width=15).place(rely=.82,relx=.60)
tk.Button(window, text="Create Account", command=create_account,width=20,bg="black",fg="white",bd=0,font=("times",15)).place(rely=.71,relx=.50)
tk.Button(window, text="Forgot Password", command=forgot_password,bg="black",fg="white",bd=0,font=("times",15)).place(rely=.71,relx=.80)

window.mainloop()







