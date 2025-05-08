import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from tkinter import PhotoImage
import os

file_path = "users.xlsx"
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Username", "Password"])
    wb.save(file_path)


wb = load_workbook(file_path)
ws = wb.active

def user_exists(username):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return True
    return False

def get_password(username):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return row[1]
    return None

def add_user(username, password):
    if not user_exists(username):
        ws.append([username, password])
        wb.save(file_path)
        return True
    return False

def login():
    username = username_entry.get()
    password = password_entry.get()
    if get_password(username) == password:
        messagebox.showinfo("Success", "Login Successful!")
    else:
        messagebox.showerror("Error", "Invalid username or password.")

def create_account():
    username = username_entry.get()
    password = password_entry.get()
    if username == "" or password == "":
        messagebox.showwarning("Warning", "Username and password required.")
    elif add_user(username, password):
        messagebox.showinfo("Success", "Account created successfully.")
    else:
        messagebox.showerror("Error", "User already exists.")

def forgot_password():
    username = username_entry.get()
    if not username:
        messagebox.showwarning("Warning", "Enter your username to retrieve password.")
    else:
        password = get_password(username)
        if password:
            messagebox.showinfo("Password Found", f"Password: {password}")
        else:
            messagebox.showerror("Error", "User not found.")
window= tk.Tk()
window.title("MIS")
window.config(bg="black")
window.geometry("850x490")

backgroundp=PhotoImage(file='login.png')
tk.Label(window,image=backgroundp).pack()

tk.Label(window, text="Username",font=("Times",25),bg="Black",fg="White").place(rely=.27,relx=.58)
username_entry = tk.Entry(window,font=("Arial",20),bd=0)
username_entry.place(height=40,width=170,rely=.37,relx=.59)

tk.Label(window, text="Password",font=("times",25),bg="black",fg="White").place(rely=.48,relx=.58)
password_entry = tk.Entry(window, show="*",font=("Times",20),bd=0)
password_entry.place(width=170,height=40,rely=.58,relx=.59)

tk.Button(window, text="Login",font=("times",25), command=login,bg="Blue",fg="white",bd=0,width=15).place(rely=.82,relx=.60)
tk.Button(window, text="Create Account", command=create_account,width=20,bg="black",fg="white",bd=0,font=("times",15)).place(rely=.71,relx=.50)
tk.Button(window, text="Forgot Password", command=forgot_password,bg="black",fg="white",bd=0,font=("times",15)).place(rely=.71,relx=.80)

window.mainloop()